from fpdf import FPDF
import streamlit as st
import pandas as pd
import io
from datetime import datetime

# ✅ Initialize invoice_date in session state
if "invoice_date" not in st.session_state:
    st.session_state.invoice_date = datetime.today().date()

file_url = "https://github.com/NGO-Xuan/Invoice-App/raw/refs/heads/main/Price%20List.xlsx"

@st.cache_data
def load_data():
    df = pd.read_excel(file_url, engine="openpyxl")
    return df

df = load_data()

if "invoice_data" not in st.session_state:
    st.session_state.invoice_data = pd.DataFrame(columns=["Brand", "NDC#", "Qty", "Expiration", "Condition", "Price", "Total"])

if "tracking_number" not in st.session_state:
    st.session_state.tracking_number = ""

# Tabs
tab1, tab2 = st.tabs(["Price Search", "Invoice"])

# ------------------------- 🔹 TAB 1: PRICE SEARCH 🔹 -------------------------
with tab1:
    st.title("\U0001F4CA Price Search Dashboard")

    brand = st.text_input("\U0001F50D Search by Brand:")
    ref_ndc = st.text_input("\U0001F50D Search by Ref# (NDC):")
    item_type = st.text_input("\U0001F50D Search by Type:")

    filtered_df = df.copy()
    if brand:
        filtered_df = filtered_df[filtered_df["Brand"].astype(str).str.contains(brand, case=False, na=False)]
    if ref_ndc:
        filtered_df = filtered_df[filtered_df["Ref# (NDC)"].astype(str).str.contains(ref_ndc, case=False, na=False)]
    if item_type:
        filtered_df = filtered_df[filtered_df["Type"].astype(str).str.contains(item_type, case=False, na=False)]

    st.write(f"Showing {len(filtered_df)} results:")

    for i, row in filtered_df.iterrows():
        cols = st.columns(len(filtered_df.columns) + 2)  # Add input box and select button
        for j, col_name in enumerate(filtered_df.columns):
            cols[j].write(row[col_name])

        qty = cols[-2].number_input("Qty", min_value=1, value=1, key=f"qty_{i}")

        if cols[-1].button("➕ Select", key=f"select_{i}"):
            selected_data = pd.DataFrame([row[["Brand", "Ref# (NDC)", "Price"]]])
            selected_data.rename(columns={"Ref# (NDC)": "NDC#"}, inplace=True)
            selected_data["Qty"] = qty  # Use selected quantity
            selected_data["Expiration"] = ""
            selected_data["Condition"] = ""
            selected_data["Total"] = selected_data["Price"] * selected_data["Qty"]
            selected_data = selected_data.fillna("")
            st.session_state.invoice_data = pd.concat([st.session_state.invoice_data, selected_data], ignore_index=True)
            st.success(f"✅ {row['Brand']} added to Invoice with Qty {qty}!")

# ------------------------- 🔹 TAB 2: INVOICE -------------------------
with tab2:
    st.title("🧾 Invoice")

    if not st.session_state.invoice_data.empty:
        st.session_state.invoice_date = st.date_input("📅 Select Invoice Date:", st.session_state.invoice_date)

        # === Calculate total and append to view (but not to the session state yet) ===
        total_invoice_sum = st.session_state.invoice_data["Qty"].astype(float) * st.session_state.invoice_data["Price"].astype(float)
        st.session_state.invoice_data["Total"] = total_invoice_sum

        total_sum = total_invoice_sum.sum()
        total_row = pd.DataFrame([{
            "Brand": "**Total Invoice**",
            "NDC#": "",
            "Qty": "",
            "Expiration": "",
            "Condition": "",
            "Price": "",
            "Total": total_sum
        }])
        invoice_with_total = pd.concat([st.session_state.invoice_data, total_row], ignore_index=True)

        # === Display the invoice table with wider layout ===
        st.dataframe(invoice_with_total, use_container_width=True)

        if st.button("🔄 Refresh Totals"):
            st.session_state.invoice_data["Total"] = st.session_state.invoice_data["Qty"].astype(float) * st.session_state.invoice_data["Price"].astype(float)
            st.success("✅ Totals updated!")

        col1, col2 = st.columns([2, 1])
        with col1:
            st.session_state.tracking_number = st.text_input("📦 Tracking #", st.session_state.tracking_number)
        with col2:
            st.write("🚚 UPS")

        st.write("\n💳 **Please Make Payment to Paypal**")
        st.write("Zelle: **derek@stripbuyer.com**")
        st.write("\n🏢 **Strip Buyer Surplus LLC**")
        st.write("2664 Alfreda Way")
        st.write("Redding, CA 96002")
        
        # === Generate PDF ===
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", style="B", size=14)
        invoice_date = st.session_state.invoice_date.strftime("%Y-%m-%d")
        pdf.cell(250, 10, "", border=0)
        pdf.cell(25, 10, f"Invoice Date: {invoice_date}", ln=True, align="R")
        pdf.cell(275, 10, "Invoice", ln=True, align="C")
        pdf.ln(10)
        pdf.set_font("Arial", size=10)
        col_names = ["Brand", "NDC#", "Qty", "Expiration", "Condition", "Price", "Total"]
        col_widths = [60, 40, 20, 35, 35, 25, 40]
        for col, width in zip(col_names, col_widths):
            pdf.cell(width, 10, col, border=1, align="C")
        pdf.ln()
        for _, row in invoice_with_total.iterrows():
            for col, width in zip(col_names, col_widths):
                pdf.cell(width, 10, str(row[col]), border=1, align="C")
            pdf.ln()
        pdf.ln(10)
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 10, f"Tracking #: {st.session_state.tracking_number}   UPS", ln=True)
        pdf.ln(5)
        pdf.cell(0, 10, "Please Make Payment to Paypal", ln=True)
        pdf.cell(0, 10, "Zelle: derek@stripbuyer.com", ln=True)
        pdf.ln(5)
        pdf.cell(0, 10, "Strip Buyer Surplus LLC", ln=True)
        pdf.cell(0, 10, "2664 Alfreda Way", ln=True)
        pdf.cell(0, 10, "Redding, CA 96002", ln=True)
        pdf_output_data = pdf.output(dest="S").encode("latin1")
        pdf_output = io.BytesIO(pdf_output_data)

        # === Generate Excel ===
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        bold_font = Font(bold=True)
        big_bold_font = Font(size=14, bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws["F1"] = "Invoice Date:"
        ws["G1"] = invoice_date
        ws["F1"].font = bold_font
        ws.merge_cells("A2:G2")
        ws["A2"] = "INVOICE"
        ws["A2"].font = big_bold_font
        ws["A2"].alignment = Alignment(horizontal="center")

        current_row = 4
        for r_idx, row in enumerate(dataframe_to_rows(invoice_with_total, index=False, header=True), start=current_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    cell.font = bold_font
                cell.border = thin_border

        current_row += len(invoice_with_total) + 2
        ws[f"A{current_row}"] = "Tracking #:"
        ws[f"B{current_row}"] = st.session_state.tracking_number
        ws[f"C{current_row}"] = "UPS"
        ws[f"A{current_row}"].font = bold_font

        current_row += 2
        ws[f"A{current_row}"] = "Please Make Payment to Paypal"
        ws[f"A{current_row + 1}"] = "Zelle: derek@stripbuyer.com"

        current_row += 3
        ws[f"A{current_row}"] = "Strip Buyer Surplus LLC"
        ws[f"A{current_row + 1}"] = "2664 Alfreda Way"
        ws[f"A{current_row + 2}"] = "Redding, CA 96002"

        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        # === Download Buttons ===
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 Download Excel",
                data=excel_output,
                file_name="invoice.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            st.download_button(
                label="📥 Download PDF",
                data=pdf_output,
                file_name="invoice.pdf",
                mime="application/pdf"
            )
